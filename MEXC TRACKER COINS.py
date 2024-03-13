from mexc_sdk import Spot
import pandas as pd
import openpyxl
import winsound
import os
import time
import tkinter as tk
import sys
from tkinter import ttk
keep_running=True

os.system('')
variable = {}

with open('Parameters_mexc.txt', 'r') as file:
    for line in file:
        name, _, value = line.partition('=')
        variable[name.strip()] = value.strip()

MEXC_API_KEY=(variable['MEXC_API_KEY'])
MEXC_API_SECRET_KEY=(variable['MEXC_API_SECRET_KEY'])
MEXC_TRACE_SYMBOL=(variable['MEXC_TRACE_SYMBOL'])
FILTRAR_USDT=(variable['FILTRAR_USDT'])
FILTRAR_BUSD=(variable['FILTRAR_BUSD'])
SONIDO=(variable['SONIDO'])
STARTTIME=(variable['STARTTIME'])
INTERVAL=(variable['INTERVAL'])
BOLLINGER_VALOR_ARRIBA=(variable['BOLLINGER_VALOR_ARRIBA'])
BOLLINGER_VALOR_ABAJO=(variable['BOLLINGER_VALOR_ABAJO'])
TIEMPO=float(variable['TIEMPO'])
RSI_UP=float(variable['RSI_UP'])
RSI_DOWN=int(variable['RSI_DOWN'])
PERIOD_RSI=int(variable['PERIOD_RSI'])

# Leemos el archivo lista de OI

try:
    data_futures = pd.read_csv('open_interest_mexc.csv')
except:
    data_futures = pd.DataFrame()


# Crear una instancia del cliente de MEXC
client = Spot()

# Definir el símbolo del par de trading (en este caso, BTCUSDT)
symbol = "PPAIUSDT"

monedas_personales = eval(MEXC_TRACE_SYMBOL)
# Definir el intervalo de tiempo (por ejemplo, 15 minutos)
interval = INTERVAL

#creamos listas blancas

listadoVelaLong = []
listadoVelaShort = []
list_RSI_up=[]
list_RSIU=[]     
list_RSI_down=[]
list_RSID=[]
line_data_futures = {}


#guardamos los precios de bitcoin y su precio actual con artificios
prices_btc_high=[]
prices_btc_low=[]
current_price_btc=2
prices_btc_high.append(3)
prices_btc_low.append(2)

# crear un DataFrame con dos filas
tracking_list = pd.DataFrame(columns=['Coin', 'var all %', 'last 6 %', 'Var 16/BTC %', '|Var all %|', '|Last 6 %|','|Var 16/BTC %|'])


#creacion de clase tabla
class Tabla(tk.Canvas):
    def __init__(self, parent, *args, **kwargs):
        tk.Canvas.__init__(self, parent, *args, **kwargs)
        self.parent = parent

        # Creamos una tabla con 4 columnas
        self.tabla = ttk.Treeview(self, columns=('Coin','var all %', 'last 6 %', 'Var 16/BTC %', '|Var all %|', '|Last 6 %|', '|Var 16/BTC %|'), show='headings', height=173)
        self.tabla.heading('Coin', text='Coin')
        self.tabla.heading('var all %', text='var all %')
        self.tabla.heading('last 6 %', text='last 6 %')
        self.tabla.heading('Var 16/BTC %', text='Var 16/BTC %')
        self.tabla.heading('|Var all %|', text='|Var all %|')
        self.tabla.heading('|Last 6 %|', text='|Last 6 %|')
        self.tabla.heading('|Var 16/BTC %|', text='|Var 16/BTC %|')

        # Configuramos el estilo de la tabla
        style = ttk.Style()
        style.configure('Treeview', rowheight=30, font=('Arial', 12), background="gray25", foreground="white")
        style.configure('Treeview.Heading', background='red', foreground='purple', font=('Helvetica', 15))
        style.map('Treeview', foreground=[('selected', 'white')])

        # Colocamos la tabla en el canvas
        self.create_window((0, 0), window=self.tabla, anchor='nw')

        # Configuramos el canvas para que se pueda desplazar con la rueda del ratón
        self.bind_all("<MouseWheel>", self.mover_tabla)

        # Inicializamos los atributos
        self.data = pd.DataFrame()
        self.num_rows = 0
        self.num_cols = 7
        self.labels = []
        self.ultima_columna_ordenada = None
        self.ultimo_orden = True
        # Establece el ancho de la columna 'Coin' en 100 unidades
        self.tabla.column('Coin', width=110)  
        self.tabla.column('var all %',  width=110)
        self.tabla.column('last 6 %',  width=110)
        self.tabla.column('Var 16/BTC %', width=140)
        self.tabla.column('|Var all %|',  width=110)
        self.tabla.column('|Last 6 %|', width=110)
        self.tabla.column('|Var 16/BTC %|',  width=110)


    def mover_tabla(self, event):
        # Movemos la tabla hacia arriba o abajo dependiendo de la dirección de la rueda del ratón
        self.yview_scroll(int(-1*(event.delta/120)), "units")
    
    def actualizar_datos(self, data):
        # Limpiamos la tabla
        self.tabla.delete(*self.tabla.get_children())

        # Añadimos los nuevos datos
        for i, row in data.iterrows():
            self.tabla.insert('', 'end', values=tuple(row))

        # Actualizamos los atributos
        self.data = data
        self.num_rows = len(data)

        
def ordenar_columna(tabla, columna, orden):
    global tracking_list
    # Si la columna que se está ordenando es la misma que la anterior, cambiamos el orden
    if columna == tabla.ultima_columna_ordenada:
        orden = not tabla.ultimo_orden
    # Ordenamos los datos según la columna y el orden especificados
    tracking_list = tracking_list.sort_values(by=columna, ascending=orden)
    # Actualizamos la tabla con los datos ordenados
    tabla.actualizar_datos(tracking_list)
    # Guardamos el estado actual de la ordenación
    tabla.ultima_columna_ordenada = columna
    tabla.ultimo_orden = orden


def agregar_datos(tabla, tracking_list):
    # Añadimos los datos del DataFrame a la tabla
    for i, row in tracking_list.iterrows():
        tabla.tabla.insert('', 'end', values=tuple(row))       

        
#text redirector

class TextRedirector(object):
    def __init__(self, widget, tag="stdout"):
        self.widget = widget
        self.tag = tag

    def write(self, str):
        self.widget.configure(state=tk.NORMAL)
        self.widget.insert("end", str, (self.tag,))
        self.widget.configure(state=tk.DISABLED)
        self.widget.see(tk.END)
        
        
def on_mousewheel(event):
    # Ajusta el valor de desplazamiento de la scrollbar en función de la dirección del desplazamiento
    if event.delta > 0:
        text_cmd.yview_scroll(-1, "units")
    else:
        text_cmd.yview_scroll(1, "units")

def get_data_frame():
    
    global symbol
    # valid intervals - 1m, 3m, 5m, 15m, 30m, 1h, 2h, 4h, 6h, 8h, 12h, 1d, 3d, 1w, 1M
    # request historical candle (or klines) data using timestamp from above, interval either every min, hr, day or month


    bars = client.klines(symbol=symbol, interval=interval)
    #print(bars)
    for line in bars:        # Keep only first 5 columns, "date" "open" "high" "low" "close"
        del line[5:]
        #del line[:4]
    df = pd.DataFrame(bars, columns=['date', 'open', 'high', 'low', 'close']) #  2 dimensional tabular data
    #print(df)
    return (df)


#calculando la variacion de cada token
def coins_price_action(symbol, symbol_df, current_price, open_interest):
    global tracking_list, prices_btc_high, prices_btc_low, current_price_btc
    
    prices_coin_high=list(map(float, symbol_df['high']))
    prices_coin_low=list(map(float, symbol_df['low']))
    
    #guardando los precios de bitcoin para el ciclo
    if symbol=="BTCUSDT":
        prices_btc_high=prices_coin_high
        prices_btc_low=prices_coin_low
        current_price_btc=current_price
        
    
   #variacion max de todas las ultimas 48 velas

    price_coin_up_all=max(prices_coin_high) #podriamos agarrar el high
    price_coin_down_all=min(prices_coin_low) #podriamos agarrar el low

    coin_var_up_amount_all=price_coin_up_all-current_price
    coin_var_down_amount_all=current_price-price_coin_down_all
    if coin_var_up_amount_all<coin_var_down_amount_all:
        coin_var_all=round(((price_coin_up_all-price_coin_down_all)*100/price_coin_down_all),2)
    else:
        coin_var_all=-round(((price_coin_up_all-price_coin_down_all)*100/price_coin_up_all),2)

    #variacion ultima hora y media 6 velas
    #capturando las ultimas 6 velas
    prices_coin_lasts_high=prices_coin_high[-6:]
    prices_coin_lasts_low=prices_coin_low[-6:]
    price_coin_up=max(prices_coin_lasts_high) 
    price_coin_down=min(prices_coin_lasts_low) 

    coin_var_up_amount=price_coin_up-current_price
    coin_var_down_amount=current_price-price_coin_down
    if coin_var_up_amount<coin_var_down_amount:
        coin_var_lasts=round(((price_coin_up-price_coin_down)*100/price_coin_down),2)
    else:
        coin_var_lasts=-round(((price_coin_up-price_coin_down)*100/price_coin_up),2)

    #var respecto a bitcoin en las ultimas 16 velas
    #obteniendo la variacion de bitcoin en las ultimas 16 velas
    up_btc=max(prices_btc_high[-16:])
    down_btc=min(prices_btc_low[-16:])
    btc_var_up_amount=up_btc-current_price_btc
    btc_var_down_amount=current_price_btc-down_btc

    #variacion de bitcoin hacia arriba
    btc_var_up=(up_btc-down_btc)*100/down_btc

    #variacion de bitcoin hacia abajo
    btc_var_down=(up_btc-down_btc)*100/up_btc

    #variacion de monedas en las ultimas 16 velas
    prices_coin_lasts_hours_high=prices_coin_high[-16:]
    prices_coin_lasts_hours_low=prices_coin_low[-16:]
    
    #sacando el maximo y el minimo precio de las ultimas 16 velas
    price_coin_up_lasts_hours=max(prices_coin_lasts_hours_high)
    price_coin_down_lasts_hours=min(prices_coin_lasts_hours_low)
    
    #calculando la longitud de la variacion con respecto al precio actual
    coin_var_up_amount_all_2=price_coin_up_lasts_hours-current_price
    coin_var_down_amount_all_2=current_price-price_coin_down_lasts_hours
    
    #si la distancia con el precio actual es mayor o menor
    if coin_var_up_amount_all_2<coin_var_down_amount_all_2:
        #variacion a la alza en las ultimas 16 velas con respecto a bitcoin
        coin_var_all_2=round(((price_coin_up_lasts_hours-price_coin_down_lasts_hours)*100/price_coin_down_lasts_hours)*100/btc_var_up,2)
    else: 
        coin_var_all_2=-round(((price_coin_up_lasts_hours-price_coin_down_lasts_hours)*100/price_coin_up_lasts_hours)*100/btc_var_down,2)
    
    #valor absoluto de los movimientos
    interest_coin_all=abs(coin_var_all)
    coin_var_6_OI=abs(coin_var_lasts)
    coin_var_all_OI=abs(coin_var_all_2)
    
    #interest abierto
    ##############################################################

    #calculando el interes abierto de la moneda en total USD
    #interest_coin_all=round((current_price*open_interest)/1000)
       
    #variacion de OI de las ultimas 6 velas
    #data_futures_last_6 = data_futures[-6:]
    
    #verificamos si la lista no esta vacia y calculamos maximos y minimos
    #try:
    #    OI_coin_up_6=max(data_futures_last_6[symbol+"_OI"])
    #    OI_coin_down_6=min(data_futures_last_6[symbol+"_OI"])
    #except:
    #    OI_coin_up_6=0
    #    OI_coin_down_6=0
            
    #calculando la longitud de la variacion con respecto al interes abierto actual
    #coin_var_up_OI_6=OI_coin_up_6-open_interest
    #coin_var_down_OI_6=open_interest-OI_coin_down_6
    
    #si la distancia con el OI actual es mayor o menor
    #evitando la division entre 0
    #if OI_coin_up_6 == 0 or OI_coin_down_6 == 0:
    #    coin_var_6_OI = 0 
    #else:
    #    if coin_var_up_OI_6 < coin_var_down_OI_6:
            # Variación al alza en las últimas 6 velas de OI
    #        coin_var_6_OI = round(((OI_coin_up_6 - OI_coin_down_6) * 100 / OI_coin_down_6), 2)
    #    else:
    #        coin_var_6_OI = -round(((OI_coin_down_6 - OI_coin_up_6) * 100 / OI_coin_up_6), 2)
    
    

    #variacion de las ultimas 16 velas
    #data_futures_last = data_futures[-16:]
    
    #verificamos si la lista no esta vacia y calculamos maximos y minimos
    #try:
    #    OI_coin_up=max(data_futures_last[symbol+"_OI"])
    #    OI_coin_down=min(data_futures_last[symbol+"_OI"])
    #except:
    #    OI_coin_up=0
    #    OI_coin_down=0
                
    #calculando la longitud de la variacion con respecto al interes abierto actual
    #coin_var_up_OI_all=OI_coin_up-open_interest
    #coin_var_down_OI_all=open_interest-OI_coin_down
    
    #si la distancia con el OI actual es mayor o menor
    #evitando la division entre 0
    #if OI_coin_up == 0 or OI_coin_down == 0:
    #    coin_var_all_OI = 0 
    #else:
    #    if coin_var_up_OI_all < coin_var_down_OI_all:
            # Variación al alza en las últimas 16 velas de OI
    #        coin_var_all_OI = round(((OI_coin_up - OI_coin_down) * 100 / OI_coin_down), 2)
    #    else:
    #        coin_var_all_OI = -round(((OI_coin_down - OI_coin_up) * 100 / OI_coin_up), 2)
            
    #anadimos los datos a la tabla
    #####################################################################################
    # actualizar la variacion de todas las velas y OI en la tabla
    if symbol in tracking_list['Coin'].values:
        tracking_list.loc[tracking_list['Coin'] == symbol, 'var all %'] = coin_var_all
    else:
        nueva_fila = {'Coin': symbol, 'var all %': coin_var_all}
        nueva_fila_df = pd.DataFrame([nueva_fila])  # Crear un DataFrame con la nueva fila
        if not nueva_fila_df.empty:
            tracking_list = pd.concat([tracking_list, nueva_fila_df], ignore_index=True)

    # actualizar la variacion de las ultimas 6 velas
    if symbol in tracking_list['Coin'].values:
        tracking_list.loc[tracking_list['Coin'] == symbol, 'last 6 %'] = coin_var_lasts
    else:
        nueva_fila = {'Coin': symbol, 'last 6 %': coin_var_lasts}
        nueva_fila_df = pd.DataFrame([nueva_fila])  # Crear un DataFrame con la nueva fila
        if not nueva_fila_df.empty:
            #nueva_fila_df = nueva_fila_df.fillna(0)
            tracking_list = pd.concat([tracking_list, nueva_fila_df], ignore_index=True)

    # actualizar la variacion de la moneda respecto a Bitcoin
    if symbol in tracking_list['Coin'].values:
        tracking_list.loc[tracking_list['Coin'] == symbol, 'Var 16/BTC %'] = coin_var_all_2
    else:
        nueva_fila = {'Coin': symbol, 'Var 16/BTC %': coin_var_all_2}
        nueva_fila_df = pd.DataFrame([nueva_fila])  # Crear un DataFrame con la nueva fila
        if not nueva_fila_df.empty:
            #nueva_fila_df = nueva_fila_df.fillna(0)
            tracking_list = pd.concat([tracking_list, nueva_fila_df], ignore_index=True)

    #actualizamos las tablas de interes abierto
    ############################################################################################
    # actualizar el interes abierto actual
    if symbol in tracking_list['Coin'].values:
        tracking_list.loc[tracking_list['Coin'] == symbol, '|Var all %|'] = interest_coin_all
    else:
        nueva_fila = {'Coin': symbol, '|Var all %|': interest_coin_all}
        nueva_fila_df = pd.DataFrame([nueva_fila])  # Crear un DataFrame con la nueva fila
        if not nueva_fila_df.empty:
            #nueva_fila_df = nueva_fila_df.fillna(0)
            tracking_list = pd.concat([tracking_list, nueva_fila_df], ignore_index=True)
        

    # actualizar el interes abierto en porcentaje total de las ultimas 6 velas
    if symbol in tracking_list['Coin'].values:
        tracking_list.loc[tracking_list['Coin'] == symbol, '|Last 6 %|'] = coin_var_6_OI
        
    else:
        nueva_fila = {'Coin': symbol, '|Last 6 %|': coin_var_6_OI}
        nueva_fila_df = pd.DataFrame([nueva_fila])  # Crear un DataFrame con la nueva fila
        if not nueva_fila_df.empty:
            #nueva_fila_df = nueva_fila_df.fillna(0)
            tracking_list = pd.concat([tracking_list, nueva_fila_df], ignore_index=True)
        
    # actualizar el interes abierto en porcentaje total
    if symbol in tracking_list['Coin'].values:
        tracking_list.loc[tracking_list['Coin'] == symbol, '|Var 16/BTC %|'] = coin_var_all_OI
    else:
        nueva_fila = {'Coin': symbol, '|Var 16/BTC %|': coin_var_all_OI}
        nueva_fila_df = pd.DataFrame([nueva_fila])  # Crear un DataFrame con la nueva fila
        if not nueva_fila_df.empty:
            #nueva_fila_df = nueva_fila_df.fillna(0)
            tracking_list = pd.concat([tracking_list, nueva_fila_df], ignore_index=True)

    #funcion para guardar los simbolos de todas las monedas
########################################################################################################
# Función para leer el archivo de texto y devolver el contenido como una cadena
def read_file_exchange_coins(file_name_exchange):
    with open(file_name_exchange, 'r') as file_coins:
        return file_coins.read()

# Función para obtener la lista actual de monedas del exchange
def extract_save_coins(current_coins_exchange):
    return [symbol['symbol'] for symbol in current_coins_exchange['symbols']]

# Función para comparar las listas y encontrar monedas desenlistadas y nuevas monedas
def compare_lists(list_file_coins_old, list_exchange):
    monedas_desenlistadas = [moneda for moneda in list_file_coins_old if moneda not in list_exchange]
    nuevas_monedas = [moneda for moneda in list_exchange if moneda not in list_file_coins_old]
    return monedas_desenlistadas, nuevas_monedas

# Función para guardar las monedas desenlistadas y nuevas monedas en archivos de texto
def save_coins_in_files(desenlistadas, nuevas, currentCryptos_exchange):
    global new_coins_exchange, new_coins_exchange_1
    
    # Guardar las monedas desenlistadas 
    with open('unlisted_coins.txt', 'a') as file:
        for moneda in desenlistadas:
            file.write(moneda + '\n')
            
    #abre y guarda en una nueva linea las monedas nuevas
    with open('new_coins.txt', 'a') as file:
        for moneda in nuevas:
            file.write(moneda + '\n')
            
    # Guardar la lista de todas kas monedas de mexc
    with open("mexc_coins.txt", 'w') as file:
        # Convertir la variable a una cadena y escribirla en el archivo
        file.write(str(currentCryptos_exchange))    
            
    #abrimos el txt con las ultimas monedas
    with open('new_coins.txt', 'r') as file:
        new_listing_coins = file.readlines()

    #extraemos las ultimas monedas y las guardamos en una variable
    new_coins_exchange = new_listing_coins[-3:]
    new_coins_exchange = [coin.strip() for coin in new_coins_exchange]
    new_coins_exchange_1 = new_listing_coins[-6:-3]
    new_coins_exchange_1 = [coin_1.strip() for coin_1 in new_coins_exchange_1]
 

#############################################################################################
    #buy sell coin
def buy_or_sell(df):
    #obtencion del precio actual de la moneda
    current_price = float(client.ticker_price(symbol=symbol)["price"])
    
    current_upper = pd.to_numeric(df.iloc[-1:]['upper'], downcast='float').iloc[0]
    current_lower = pd.to_numeric(df.iloc[-1:]['lower'], downcast='float').iloc[0]
    #upper_lower_delta = current_upper - current_lower
    #current_price_delta = current_price - current_lower
    #current_price_percentage = current_price_delta / upper_lower_delta * 100.0
    percentage_current_upper = ((current_price - current_upper) / current_upper) * 100
    percentage_current_lower = ((current_lower - current_price) / current_lower) * 100

    return (current_price, current_upper, current_lower, percentage_current_upper, percentage_current_lower, symbol)

#Bollinger Data and RSI

def bollinger_trade_logic():
    symbol_df = get_data_frame()
    period = 20
    
    #rsi
    symbol_df.close=pd.to_numeric(symbol_df.close)
    chg=symbol_df.close.diff(1)
    gain=chg.mask(chg<0,0)
    symbol_df['gain']=gain
    
    loss=chg.mask(chg>0,0)
    symbol_df['loss']=loss
    
    avg_gain=gain.ewm(com=PERIOD_RSI-1, min_periods= PERIOD_RSI).mean()
    avg_loss=loss.ewm(com=PERIOD_RSI-1, min_periods= PERIOD_RSI).mean()
    
    symbol_df['avg gain'] = avg_gain
    symbol_df['avg loss'] = avg_loss
    
    rs= abs(avg_gain/avg_loss)
    rsi= round(100-(100/(1+rs)))
    symbol_df['rsi']=pd.to_numeric(rsi)
    rsi_token=rsi.tail(1)

    
    # small time Moving average. calculate 20 moving average using Pandas over close price
    symbol_df['sma'] = symbol_df['close'].rolling(period).mean()
    # Get standard deviation
    symbol_df['std'] = symbol_df['close'].rolling(period).std()
    # Calculate Upper Bollinger band
    symbol_df['upper'] = symbol_df['sma']  + (2 * symbol_df['std'])
    # Calculate Lower Bollinger band
    symbol_df['lower'] = symbol_df['sma']  - (2 * symbol_df['std'])
    # To print in human readable date and time (from timestamp)
    symbol_df.set_index('date', inplace=True)
    symbol_df.index = pd.to_datetime(symbol_df.index, unit='ms') # index set to first column = date_and_time
   
    (current_price, current_upper, current_lower, percentage_current_upper, percentage_current_lower, symbol) = buy_or_sell(symbol_df)
    return (current_price, current_upper, current_lower, percentage_current_upper, percentage_current_lower, symbol, rsi_token, symbol_df)

def mainloop():
    global symbol, current_price, prevLen, prevLen_f, trading_pairs, last_coin, last_coin_f, currentLen_f, prevLen_f, prevLen, currentLen, ciclo_bloque, cantidad_pares
    #futuros


    # Nombre del archivo que contiene la lista de monedas
    list_coins_exchange = 'mexc_coins.txt'
    
    current_coins_exchange = client.exchange_info()
    currentCryptos_f=current_coins_exchange["symbols"]
    
    # Leer la lista de monedas del archivo
    list_file_coins_old = eval(read_file_exchange_coins(list_coins_exchange))
    
    # Obtener la lista actual de monedas del exchange
    currentCryptos_exchange = extract_save_coins(current_coins_exchange)
    
    # Comparar las listas para encontrar monedas desenlistadas y nuevas monedas
    monedas_desenlistadas, nuevas_monedas = compare_lists(list_file_coins_old, currentCryptos_exchange)
    
    # Guardar las monedas desenlistadas y nuevas monedas en archivos de texto
    save_coins_in_files(monedas_desenlistadas, nuevas_monedas, currentCryptos_exchange)
    
    #contamos cuantas monedas tiene
    currentLen_f = len(currentCryptos_f)
    
    #impresion de los datos de cmd
    
    if prevLen_f < currentLen_f:
        print ("new Futures coin")
        newCoin_f = prevLen_f
        for i in range (newCoin_f , currentLen_f):
            amount_f = 0
            last_coin_f = currentCryptos_f[i].get('symbol')
        print(last_coin_f)
        winsound.PlaySound('new_coin.wav', winsound.SND_ALIAS)
        prevLen_f=currentLen_f
    else:
        last_coin_f=currentCryptos_f[prevLen_f-1].get('symbol')
        
    
    #Spot
    #currentCryptos = currentCryptos_f
    #currentLen = currentLen_f

    #if prevLen < currentLen:
    #    print ("new Spot coin")
    #    newCoin = prevLen
    #    for i in range (newCoin , currentLen):
    #        amount = 0
    #        last_coin = currentCryptos[i].get('symbol')
    #    print(last_coin)
    #    winsound.PlaySound('new_coin.wav', winsound.SND_ALIAS)
    #    prevLen=currentLen
    #else:
    #    last_coin=currentCryptos[prevLen-1].get('symbol')
    
    most_recently_coins=("Last Coins Spot: \n" + str(new_coins_exchange) + "\n" + str(new_coins_exchange_1))
    text_cmd_2.config(text=most_recently_coins)
    
    #Determinar si busca monedas personales o las monedas de MEXC
    ciclo_bloque = 0
    if len(monedas_personales) == 0:
        # request info on all futures symbols
        trading_pairs = [info['symbol'] for info in currentCryptos_f] #Estrae todas las monedas de futuros de MEXC
        if FILTRAR_USDT != "SI" and FILTRAR_BUSD != "SI":
            pares = trading_pairs
            cantidad_pares = len(pares)
        elif FILTRAR_USDT == "SI" and FILTRAR_BUSD != "SI":
            pares = [data for data in trading_pairs if "USDT" in data] #filtra solo las monedas USDT
            cantidad_pares = len(pares)
        elif FILTRAR_USDT != "SI" and FILTRAR_BUSD == "SI":
            pares = [data for data in trading_pairs if "BUSD" in data] #filtra solo las monedas USDT
            cantidad_pares = len(pares)
        elif FILTRAR_USDT == "SI" and FILTRAR_BUSD == "SI":
            print("si selecciona SI en USDT y BUSD, no tendria ninguna lista.\nY me voy a repetir hasta el cansancio")
            pares = []
            cantidad_pares = len(pares)
            
        
    else:
        pares = monedas_personales
        cantidad_pares = len(pares)
   
    trading_pairs = pares 
    #finalizacion del argumento de menor frecuencia
    
    # loop para que vaya por todas las monedas

    coin_analisis_all(trading_pairs)
    
    #artificio para cerrar el bucle al finalizar todas las monedas
    if keep_running==True:
        root.after(1000,mainloop)

def coin_analisis_all(trading_pairs, index=0):
    global keep_running, symbol, last_coin, last_coin_f, currentLen_f, prevLen_f, prevLen, currentLen, ciclo_bloque, cantidad_pares, tracking_list
    keep_running=False
    #verificar si el ciclo es mayor que el conteo de monedas para salirse del bucle e ir al bucle mayor
    if index >= len(trading_pairs):
        keep_running=True
        listadoVelaLong.clear()
        listadoVelaShort.clear()
        list_RSID.clear()
        list_RSIU.clear()
        save_data_futures()
        #guardando data futures

        mainloop()
        return
    
    line=trading_pairs[index]

    #trading_pairs = ["ada", "btc", "eth", "sand", 5, 5, 8, 6, 1, 8]
    #aqui iria el argumento que recorre todas las monedas
    
    try:
        symbol = line
        (current_price, current_upper, current_lower, percentage_current_upper, percentage_current_lower, symbol, rsi_token, symbol_df) = bollinger_trade_logic()
        #analizando el Open Interest
        
        open_interest=1
        
        #analizando el % de variacion de monedas y el interes abierto
        
        coins_price_action(symbol, symbol_df, current_price, open_interest)

        if float(rsi_token.iloc[0]) >= RSI_UP:
        #if float(rsi_token)>=RSI_UP:
            list_RSI_up=str(symbol)+": "+rsi_token.to_string(index=False)#str(symbol)+": "+str(rsi_token)
            list_RSIU.append(list_RSI_up)
            winsound.PlaySound('RSI Alert.wav', winsound.SND_ALIAS)
            
        if float(rsi_token.iloc[0]) <= RSI_DOWN:
        #if float(rsi_token)<=RSI_DOWN:
            list_RSI_down=str(symbol)+": "+rsi_token.to_string(index=False)#str(symbol)+": "+str(rsi_token)
            list_RSID.append(list_RSI_down)
            winsound.PlaySound('RSI Alert.wav', winsound.SND_ALIAS)

        if percentage_current_lower >= float(BOLLINGER_VALOR_ABAJO):
            listadoVelaLong.append(symbol)
            if SONIDO == "SI":
                winsound.PlaySound('Alertcoin.wav', winsound.SND_ALIAS)
                try:
                    date_bollinger=time.strftime('%d/%m/%y', time.localtime())
                    time_bollinger=time.strftime('%H:%M:%S', time.localtime())
                    abajo="DOWN"
                    process_data(date_bollinger, time_bollinger, symbol, abajo, current_price)

                    # Abre el archivo de Excel en modo lectura/escritura
                    wb = openpyxl.load_workbook("Bollinger_data_mexc.xlsx")
                    # Selecciona la primera hoja de cálculo del libro
                    sheet = wb[wb.sheetnames[0]]
                    # Añade los datos al final de la hoja de cálculo
                    sheet.append([date1, time1, data1, status, price_now1])
                    # Guarda el archivo de Excel
                    wb.save("Bollinger_data_mexc.xlsx")
                except:
                    None

        if percentage_current_upper >= float(BOLLINGER_VALOR_ARRIBA):
            listadoVelaShort.append(symbol)
            if SONIDO == "SI":
                winsound.PlaySound('Alertcoin.wav', winsound.SND_ALIAS)
                try:
                    date_bollinger=time.strftime('%d/%m/%y', time.localtime())
                    time_bollinger=time.strftime('%H:%M:%S', time.localtime())
                    arriba="UP"
                    process_data(date_bollinger, time_bollinger, symbol, arriba, current_price)

                    # Abre el archivo de Excel en modo lectura/escritura
                    wb = openpyxl.load_workbook("Bollinger_data_mexc.xlsx")
                    # Selecciona la primera hoja de cálculo del libro
                    sheet = wb[wb.sheetnames[0]]
                    # Añade los datos al final de la hoja de cálculo
                    sheet.append([date1, time1, data1, status, price_now1])
                    # Guarda el archivo de Excel
                    wb.save("Bollinger_data_mexc.xlsx")
                except:
                    None


        
        # A partir de aquí, todas las impresiones y mensajes

        #neva moneda
        if prevLen_f != currentLen_f:
            print("!New Spot Coin!: ",last_coin_f)

        #if prevLen != currentLen:
        #    print("!New Future Coin!: ",last_coin)

        print("..............................................")
        
        #bandas de bollinger

        print(f"        Moneda: {symbol}")
        #print(f"  Temporalidad: {interval}")
        #print(f"Bollinger high: {percentage_current_upper:.2f} ({float(BOLLINGER_VALOR_ARRIBA)})")
        #print(f" Bollinger low: {percentage_current_lower:.2f} ({float(BOLLINGER_VALOR_ABAJO)})")
        print(f" Lista UPPER: {listadoVelaShort[-10:]}")
        print(f"  Lista LOWER: {listadoVelaLong[-10:]}\n")
        text_cmd.insert(tk.END, "wakaka", ("blue",))
        print(f"rsi high: {list_RSIU[-10:]}") 
        print(f"rsi low: {list_RSID[-10:]}")
       
        
        print(f"---- {(ciclo_bloque):03d} de {cantidad_pares} monedas ----")

        ciclo_bloque += 1
              
        cantidad_paresA=cantidad_pares
        tabla.actualizar_datos(tracking_list)

    except:
        pass
    window.after(4000, coin_analisis_all, trading_pairs, index+1)
    
    


#######################################
#tkinter interfaz (desde aqui inicia todo)       
    
window = tk.Tk()
window.title('MEXC Spot Tracker')
window.geometry("800x1010")


# Creamos la tabla en el centro de la ventana
marco = tk.Frame(window)
marco.grid(row=0, column=0, sticky="nsew")

tabla = Tabla(marco, width=1000, height=700, scrollregion=(0,0,800,100000))
tabla.grid(row=0, column=0, sticky="nsew")

# Creamos un marco para contener los widgets de texto
text_frame = tk.Frame(marco)
text_frame.grid(row=1, column=0, sticky="nsew")

text_cmd = tk.Text(text_frame, width=50, height=18)
text_cmd.grid(row=0, column=0, sticky="nsew")

text_cmd_2 = tk.Label(text_frame, text="¡LAST COINS!")
text_cmd_2.grid(row=0, column=1, sticky="nsew")
text_cmd_2.config(font=("Calibri", 16), bg="black", fg="white", width=36, height=3)

# crea el tag "large_font" y establece el tamaño de fuente
text_cmd.configure(font=("Calibri", 12), foreground="white", background="black")

# crea el tag "color_red" y establece el color de texto y fondo
text_cmd.tag_configure("color_red", foreground="black", background="blue")

# Vincula el evento <MouseWheel> a la función on_mousewheel
text_cmd.bind("<MouseWheel>", on_mousewheel)
sys.stdout = TextRedirector(text_cmd, "stdout")


# Crea el widget Scrollbar y vincúlalo al Text
scrollbar = tk.Scrollbar(marco, command=text_cmd.yview)
scrollbar.grid(row=1, column=1, sticky='nse')

# Configura el Text para que utilice el Scrollbar
text_cmd.config(yscrollcommand=scrollbar.set)

# Asignamos la función a cada encabezado de columna

tabla.tabla.heading('Coin', text='Coin', command=lambda: ordenar_columna(tabla, 'Coin', not tabla.ultimo_orden if tabla.ultima_columna_ordenada == 'Coin' else True))
tabla.tabla.heading('var all %', text='var all %', command=lambda: ordenar_columna(tabla, 'var all %', not tabla.ultimo_orden if tabla.ultima_columna_ordenada == 'var all' else True))
tabla.tabla.heading('last 6 %', text='last 6 %', command=lambda: ordenar_columna(tabla, 'last 6 %', not tabla.ultimo_orden if tabla.ultima_columna_ordenada == 'last 6 %' else True))
tabla.tabla.heading('Var 16/BTC %', text='Var 16/BTC %', command=lambda: ordenar_columna(tabla, 'Var 16/BTC %', not tabla.ultimo_orden if tabla.ultima_columna_ordenada == 'Var 16/BTC %' else True))
tabla.tabla.heading('|Var all %|', text='|Var all %|', command=lambda: ordenar_columna(tabla, '|Var all %|', not tabla.ultimo_orden if tabla.ultima_columna_ordenada == '|Var all %|' else True))
tabla.tabla.heading('|Last 6 %|', text='|Last 6 %|', command=lambda: ordenar_columna(tabla, '|Last 6 %|', not tabla.ultimo_orden if tabla.ultima_columna_ordenada == '|Last 6 %|' else True))
tabla.tabla.heading('|Var 16/BTC %|', text='|Var 16/BTC %|', command=lambda: ordenar_columna(tabla, '|Var 16/BTC %|', not tabla.ultimo_orden if tabla.ultima_columna_ordenada == '|Var 16/BTC %|' else True))

          
# Llamar al método klines para obtener datos históricos
# Puedes configurar opciones adicionales, como startTime, endTime y limit
# Aquí, usamos las opciones predeterminadas con limit=500

prevLen = int((variable['SPOT_COINS']))
prevLen_f = int((variable['FUTURES_COIN']))


#loop infinito
window.after(3000,mainloop)
#creando la tabla
agregar_datos(tabla, tracking_list)
window.geometry("+-1920+0")
window.attributes('-topmost',True)
window.mainloop()
